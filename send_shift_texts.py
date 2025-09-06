#!/usr/bin/env python3
"""
Free sender: Gmail SMTP (App Password) -> carrier email gateways (SMS/MMS).
- Prefers SMS for AT&T (mms.att.net often bounces)
- Optional global FORCE_SMS via .env
- Throttles between sends to reduce blocking

Usage:
  python send_shift_texts.py --dry-run
  python send_shift_texts.py
  python send_shift_texts.py --calendar "COOP_Calendar_2025-09.xlsx" --only "Alice Johnson,Ben Smith"

Requires:
  pip install pandas openpyxl python-dotenv
  Gmail account with 2FA + App Password
"""

import os, re, ssl, smtplib, time
from typing import Dict, List, Tuple, Optional
from datetime import date
from email.message import EmailMessage
from pathlib import Path
import math
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv

# ----- Load .env next to this file (works no matter where you run from) -----
load_dotenv(dotenv_path=Path(__file__).with_name(".env"), override=True)

# ----------------- Env / Config -----------------
ROSTER_XLSX = os.getenv("ROSTER_XLSX", "Employee_Schedule.xlsx")
CALENDAR_XLSX_DEFAULT: Optional[str] = os.getenv("CALENDAR_XLSX") or None

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT_SSL = 465
SMTP_USER = os.getenv("GMAIL_USER", "")
SMTP_PASS = os.getenv("GMAIL_APP_PASSWORD", "")
FROM_NAME = os.getenv("FROM_NAME", "CO-OP Scheduler")
MAX_SMS_CHARS = int(os.getenv("MAX_SMS_CHARS", "300"))

FORCE_SMS = os.getenv("FORCE_SMS", "false").strip().lower() == "true"
RATE_LIMIT_SECONDS = float(os.getenv("RATE_LIMIT_SECONDS", "2"))

if not SMTP_USER or not SMTP_PASS:
    print("[!] Missing GMAIL_USER or GMAIL_APP_PASSWORD in .env")
    # allow --dry-run to still work

# Left-hand labels in your calendar layout
PERIOD_LABELS = ["Period 1/5", "Period 2/6", "Period 3/7 Kiosk", "Period 4/8"]
PERIOD_TIMES = {
    "Period 1/5": "P1/5",
    "Period 2/6": "P2/6",
    "Period 3/7 Kiosk": "P3/7 Kiosk",
    "Period 4/8": "P4/8",
}

# Gateways — MMS (longer) + SMS fallback (more reliable)
CARRIER_MMS: Dict[str, str] = {
    "att": "mms.att.net", "at&t": "mms.att.net",
    "verizon": "vzwpix.com",
    "t-mobile": "tmomail.net", "tmobile": "tmomail.net", "t mobile": "tmomail.net",
    "sprint": "pm.sprint.com",
    "boost": "myboostmobile.com", "boost mobile": "myboostmobile.com",
    "cricket": "mms.cricketwireless.net",
    "metropcs": "mymetropcs.com", "metro": "mymetropcs.com",
    "us cellular": "mms.uscc.net",
    "google fi": "msg.fi.google.com",
    "consumer cellular": "mailmymobile.net",
    "ting": "message.ting.com",
    "xfinity mobile": "vzwpix.com",
    "visible": "vzwpix.com",
}
CARRIER_SMS_FALLBACK: Dict[str, str] = {
    "att": "txt.att.net", "at&t": "txt.att.net",
    "verizon": "vtext.com",
    "t-mobile": "tmomail.net", "tmobile": "tmomail.net", "t mobile": "tmomail.net",  # TMO uses same
    "sprint": "messaging.sprintpcs.com",
    "boost": "sms.myboostmobile.com", "boost mobile": "sms.myboostmobile.com",
    "us cellular": "email.uscc.net",
    "cricket": "sms.cricketwireless.net",
}

# ----------------- Helpers -----------------
def normalize_carrier(name: str) -> str:
    if not isinstance(name, str):
        return ""
    return re.sub(r"\s+", " ", name.strip().lower())

def normalize_phone(num: str) -> Optional[str]:
    if not isinstance(num, str):
        num = str(num)
    digits = re.sub(r"\D", "", num)
    return digits if len(digits) == 10 else None

def gateway_addr(phone: str, carrier: str) -> Tuple[Optional[str], Optional[str]]:
    ph = normalize_phone(phone)
    key = normalize_carrier(carrier)
    if not ph:
        return None, None
    mms = CARRIER_MMS.get(key)
    sms = CARRIER_SMS_FALLBACK.get(key)
    return (f"{ph}@{mms}" if mms else None, f"{ph}@{sms}" if sms else None)

def get_ws_for_today(path: Optional[str]):
    today = date.today()
    cal = path or CALENDAR_XLSX_DEFAULT or f"COOP_Calendar_{today.year:04d}-{today.month:02d}.xlsx"
    wb = load_workbook(cal)
    return wb.active, today

def find_col_for_day(ws, d: date) -> Optional[int]:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=8):
        for c in row:
            if isinstance(c.value, int) and c.value == d.day:
                return c.column
    return None

def find_row_with_label(ws, start_row: int, label: str) -> Optional[int]:
    for r in range(start_row, min(ws.max_row, start_row + 20) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label:
            return r
    return None

def parse_cell(text: str) -> Tuple[List[str], List[str]]:
    prim, alt = [], []
    if not text:
        return prim, alt
    for line in str(text).splitlines():
        t = line.strip()
        if not t:
            continue
        if t.lower().startswith("alt:"):
            alt.append(t.split(":", 1)[1].strip())
        else:
            prim.append(t)
    return prim, alt

def make_sms(name: str, items: List[Tuple[str, List[str]]], is_alt_only: bool) -> str:
    first = name.split()[0] if name else "Hi"
    head = f"{first}, you're the CO-OP alternate today." if is_alt_only else f"{first}, you're scheduled for CO-OP today."
    parts = [head]
    for label, mates in items:
        when = PERIOD_TIMES.get(label, label)
        parts.append(
            f"{when}: ALT. Primaries: {', '.join(mates) if mates else 'TBD'}"
            if is_alt_only else
            (f"{when}: with {', '.join(mates)}" if mates else f"{when}: you're on")
        )
    msg = " ".join(parts)
    return (msg[:MAX_SMS_CHARS-3] + "...") if len(msg) > MAX_SMS_CHARS else msg

# ----------------- SMTP send -----------------
def smtp_send_text(to_addr: str, body: str, subject: str = ""):
    if not SMTP_USER or not SMTP_PASS:
        raise RuntimeError("SMTP not configured: set GMAIL_USER and GMAIL_APP_PASSWORD in .env")
    msg = EmailMessage()
    msg["From"] = f"{FROM_NAME} <{SMTP_USER}>"
    msg["To"] = to_addr
    # Keep Subject blank; some gateways act weird with subjects
    if subject:
        msg["Subject"] = subject
    msg.set_content(body)

    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT_SSL, context=ctx) as smtp:
        smtp.login(SMTP_USER, SMTP_PASS)
        smtp.send_message(msg)

# ----------------- Main -----------------
def main():
    import argparse
    p = argparse.ArgumentParser(description="Text CO-OP shifts today via Gmail SMTP + carrier gateways (free).")
    p.add_argument("--calendar", default=None, help="Path to month calendar .xlsx (defaults to COOP_Calendar_YYYY-MM.xlsx)")
    p.add_argument("--dry-run", action="store_true", help="Print what would be sent without sending")
    p.add_argument("--only", default=None, help='Comma-separated names to send (filter), e.g. "Alice,Ben"')
    args = p.parse_args()

    # Roster (Name | Schedule | Phone | Carrier)
    df = pd.read_excel(ROSTER_XLSX)
    required = {"Name", "Schedule", "Phone", "Carrier"}
    if not required.issubset(df.columns):
        raise ValueError(f"{ROSTER_XLSX} must include columns: {', '.join(sorted(required))}")

    phone = {str(r["Name"]).strip(): str(r["Phone"]).strip() for _, r in df.iterrows()}
    carrier = {str(r["Name"]).strip(): str(r["Carrier"]).strip() for _, r in df.iterrows()}

    # Calendar for today
    ws, today = get_ws_for_today(args.calendar)
    col = find_col_for_day(ws, today)
    if not col:
        print(f"[!] No column for {today}. Ensure the calendar file for this month exists.")
        return

    # locate date row
    date_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == today.day:
            date_row = r
            break
    if not date_row:
        print("[!] Could not find the row containing today's date number.")
        return
    block_start = date_row + 1

    # Build assignments per person
    assignments: Dict[str, Dict] = {}
    def ensure(name: str):
        if name not in assignments:
            assignments[name] = {"primary": [], "alt": [], "mates": {}}

    for label in PERIOD_LABELS:
        row = find_row_with_label(ws, block_start, label)
        if not row:
            continue
        prim, alt = parse_cell(ws.cell(row=row, column=col).value)
        for p_ in prim:
            ensure(p_); assignments[p_]["primary"].append(label)
        for p_ in prim:
            mates = [x for x in prim if x != p_]
            assignments[p_]["mates"].setdefault(label, mates)
        for a_ in alt:
            ensure(a_); assignments[a_]["alt"].append(label)
            assignments[a_]["mates"].setdefault(label, prim)

    only_set = {x.strip() for x in args.only.split(",")} if args.only else None

    # Send
    sent, skipped = 0, 0

    for person, info in assignments.items():
        if only_set and person not in only_set:
            continue

        ph = phone.get(person, "")
        cr = carrier.get(person, "")
        addr_mms, addr_sms = gateway_addr(ph, cr)

        if not addr_mms and not addr_sms:
            print(f"[skip] {person}: invalid phone/carrier (phone must be 10 digits)")
            skipped += 1
            continue

        # Prefer SMS for AT&T (mms.att.net bounces often), or globally if FORCE_SMS
        cr_key = normalize_carrier(cr)
        prefer_sms = FORCE_SMS or cr_key in ("att", "at&t")
        target = (addr_sms or addr_mms) if prefer_sms else (addr_mms or addr_sms)
        fallback = (addr_mms if prefer_sms else addr_sms)

        is_alt_only = bool(info["alt"]) and not info["primary"]
        ordered = [lbl for lbl in PERIOD_LABELS if lbl in info["primary"] or lbl in info["alt"]]
        items = [(lbl, info["mates"].get(lbl, [])) for lbl in ordered]
        body = make_sms(person, items, is_alt_only=is_alt_only)

        if args.dry_run:
            print(f"[DRY RUN] {person} -> {target}" + (f" (fallback {fallback})" if fallback else "") + f" | {body}")
            sent += 1
            continue

        try:
            smtp_send_text(target, body, subject="")  # subject intentionally blank
        except Exception as e:
            # try fallback domain if available
            if fallback:
                try:
                    smtp_send_text(fallback, body, subject="")
                except Exception as e2:
                    print(f"[ERROR] {person}: {e2}")
                    skipped += 1
                    continue
            else:
                print(f"[ERROR] {person}: {e}")
                skipped += 1
                continue

        print(f"[sent] {person} -> {target}")
        sent += 1

        # Throttle to avoid gateway rate limits
        time.sleep(RATE_LIMIT_SECONDS)

    print(f"Done. Texts sent/queued: {sent}. Skipped/failed: {skipped}.")

if __name__ == "__main__":
    main()
