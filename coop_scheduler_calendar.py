# coop_scheduler_calendar.py

import argparse
import calendar
import math
import random
import re
from collections import defaultdict
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -------------------------
# Config / Eligibility
# -------------------------
ELIGIBLE_KEYS = ["off", "career prep", "business management", "bus mgt", "management"]

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())

def is_eligible(subject: str) -> bool:
    s = norm(subject)
    return any(k in s for k in ELIGIBLE_KEYS)

def parse_schedule_str(schedule_str: str) -> dict:
    """'1 - career prep, 2 - algebra, 3 - off, ...' -> {1:'career prep', 2:'algebra', 3:'off', ...}"""
    periods = {}
    if not isinstance(schedule_str, str) or not schedule_str.strip():
        return periods
    for entry in schedule_str.split(","):
        entry = entry.strip()
        if not entry:
            continue
        m = re.match(r"^\s*(\d+)\s*[-:]\s*(.+?)\s*$", entry)
        if not m:
            continue
        p = int(m.group(1))
        subj = m.group(2).strip()
        periods[p] = subj
    return periods

def load_students(path: str) -> dict:
    df = pd.read_excel(path)
    cols = [str(c) for c in df.columns]
    if "Name" not in cols or "Schedule" not in cols:
        raise ValueError(f"Expected columns ['Name','Schedule'], got {cols}")
    students = {}
    for _, row in df.iterrows():
        name = str(row["Name"]).strip()
        sched = parse_schedule_str(row["Schedule"])
        eligible = [p for p, subj in sched.items() if is_eligible(subj)]
        off_p = [p for p, subj in sched.items() if "off" in norm(subj)]
        cp_p  = [p for p, subj in sched.items() if "career prep" in norm(subj)]
        students[name] = {
            "eligible": set(eligible),
            "off": set(off_p),
            "cp": set(cp_p),
            "last_worked": None,
            "total": 0,
            "per_period": defaultdict(int),  # counts per period for fairness
        }
    return students


# -------------------------
# A/B mapping with anchor
# -------------------------
def ab_map_for_month_with_anchor(year: int, month: int, anchor_date: date, anchor_day: str):
    """
    Build {date: 'A'|'B'|None} for all days in the month.
    - Only weekdays (Mon-Fri) get A/B.
    - Weekends are None.
    - A/B alternation anchored so that the most recent weekday <= anchor_date
      inside the month equals anchor_day.
    """
    assert anchor_day in ("A", "B")
    first = date(year, month, 1)
    _, ndays = calendar.monthrange(year, month)
    all_days = [first + timedelta(days=i) for i in range(ndays)]

    weekdays = [d for d in all_days if d.weekday() < 5]

    # Anchor index (most recent weekday on/before anchor_date)
    if anchor_date < first:
        anchor_idx = 0
    elif anchor_date > all_days[-1]:
        anchor_idx = len(weekdays) - 1
    else:
        prev_weekdays = [d for d in weekdays if d <= anchor_date]
        anchor_idx = (len(prev_weekdays) - 1) if prev_weekdays else 0

    # If anchor_idx is even, first weekday must be anchor_day; else the opposite
    start_flag = anchor_day if (anchor_idx % 2 == 0) else ("B" if anchor_day == "A" else "A")

    ab_map = {}
    flag = start_flag
    for d in weekdays:
        ab_map[d] = flag
        flag = "B" if flag == "A" else "A"

    # Mark weekends as None
    for d in all_days:
        if d.weekday() >= 5:
            ab_map[d] = None
        elif d not in ab_map:
            ab_map[d] = None
    return ab_map


# -------------------------
# Month helpers
# -------------------------
def month_weeks_sun_to_sat(year: int, month: int):
    """Return list of weeks; each week is a list of 7 date objects (Sun..Sat), incl. spillover."""
    cal = calendar.Calendar(firstweekday=6)  # Sunday=6
    return cal.monthdatescalendar(year, month)

def weekdays_in_month(year: int, month: int):
    first = date(year, month, 1)
    _, ndays = calendar.monthrange(year, month)
    return [first + timedelta(days=i) for i in range(ndays) if (first + timedelta(days=i)).weekday() < 5]


# -------------------------
# Fair picker (3 + Alt) with stronger evenness
# -------------------------
def pick_for_period(students: dict, current_date: date, period: int,
                    assigned_today: set, target_per_student: int,
                    min_gap_days=2, need=3, enforce_one_per_day=True, seed=42):
    """
    Select 3 primaries + 1 alternate with fairness:
      - Prefer students below target_per_student
      - Enforce one-per-day (soft; relaxed if needed)
      - Prefer lower per-period count
      - Keep your priority (OFF+CP > OFF > CP > other eligible)
      - Respect min_gap_days unless period is OFF
    """
    rng = random.Random(seed + period + current_date.toordinal())

    # Build candidate universe under gap/eligibility
    universe = []
    for name, data in students.items():
        if period not in data["eligible"]:
            continue
        last = data["last_worked"]
        days_since = (current_date - last).days if last else 10**9
        if days_since >= min_gap_days or period in data["off"]:
            n_off = period in data["off"]
            n_cp  = period in data["cp"]
            if n_off and n_cp:
                pr = 4
            elif n_off:
                pr = 3
            elif n_cp:
                pr = 2
            else:
                pr = 1
            below_target = 1 if data["total"] < target_per_student else 0
            gap = max(0, target_per_student - data["total"])
            per_period_count = data["per_period"][period]
            ds = days_since if days_since != float("inf") else 9999
            universe.append({
                "name": name,
                "priority": pr,
                "below": below_target,
                "gap": gap,
                "per_period": per_period_count,
                "days_since": ds,
                "total": data["total"],
            })

    # Helper to sort candidates (higher is better, so we sort with negatives where needed)
    def sort_key(c):
        # Strong fairness first (below target, then how far below),
        # then priority, then recency, then less-used for this period, then overall totals.
        return (
            -c["below"],          # prefer below target
            -c["gap"],            # further below target first
            -c["priority"],       # higher priority first
            -c["days_since"],     # longer since worked
            c["per_period"],      # fewer at this period
            c["total"],           # fewer total shifts overall
            rng.random(),         # tiny jitter to avoid deterministic ties
        )

    # Stage 1: enforce one-per-day
    pool_strict = [c for c in universe if c["name"] not in assigned_today]
    rng.shuffle(pool_strict)
    pool_strict.sort(key=sort_key)

    selected = []
    used_names = set()
    for c in pool_strict:
        if len(selected) >= need:
            break
        if c["name"] in used_names:
            continue
        selected.append(c)
        used_names.add(c["name"])

    # Stage 2: if we still need more, relax one-per-day
    if len(selected) < need:
        pool_relaxed = [c for c in universe if c["name"] not in used_names]
        rng.shuffle(pool_relaxed)
        pool_relaxed.sort(key=sort_key)
        for c in pool_relaxed:
            if len(selected) >= need:
                break
            selected.append(c)
            used_names.add(c["name"])

    primaries = [c["name"] for c in selected]

    # Alternate: best remaining distinct from primaries; prefer not assigned earlier today
    remaining = [c for c in universe if c["name"] not in used_names]
    rng.shuffle(remaining)
    remaining.sort(key=sort_key)
    alt = None
    for c in remaining:
        if c["name"] not in assigned_today:
            alt = c["name"]
            break
    if alt is None and remaining:
        alt = remaining[0]["name"]

    # Write back only for primaries
    for nm in primaries:
        students[nm]["last_worked"] = current_date
        students[nm]["total"] += 1
        students[nm]["per_period"][period] += 1
        assigned_today.add(nm)

    return primaries, alt


# -------------------------
# Excel builder
# -------------------------
def build_calendar_excel(out_path: str, year: int, month: int, students: dict,
                         min_gap_days=2, seed=42, anchor_date: date = None, anchor_day: str = "B",
                         enforce_one_per_day=True):
    # Styles
    fill_month_title = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="A9D18E")
    fill_pink = PatternFill("solid", fgColor="F4C2C2")
    fill_day_a = PatternFill("solid", fgColor="E5DFEC")  # purple-ish
    fill_day_b = PatternFill("solid", fgColor="FFF2CC")  # yellow-ish
    fill_left = PatternFill("solid", fgColor="9BC2E6")
    fill_grey = PatternFill("solid", fgColor="DDDDDD")
    thin = Side(style="thin", color="000000")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = calendar.month_name[month]

    # Column widths: A label + B..H (Sun..Sat) + K/L legend
    ws.column_dimensions["A"].width = 22
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 16

    # Row indices
    ROW_TITLE = 1
    ROW_DOW = 2  # Day-of-week header on row 2

    # Month title across B..H (row 1)
    month_name = f"{calendar.month_name[month]} {year}"
    ws.merge_cells(start_row=ROW_TITLE, start_column=2, end_row=ROW_TITLE, end_column=8)
    c = ws.cell(row=ROW_TITLE, column=2, value=month_name)
    c.fill = fill_month_title
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ROW_TITLE].height = 22

    # Legend in K1/L1 (row 1)
    key_a = ws.cell(row=ROW_TITLE, column=11, value="A day")   # K1
    key_a.fill = fill_day_a
    key_a.alignment = Alignment(horizontal="center")
    key_a.border = border_all

    key_b = ws.cell(row=ROW_TITLE, column=12, value="B day")   # L1
    key_b.fill = fill_day_b
    key_b.alignment = Alignment(horizontal="center")
    key_b.border = border_all

    # Day-of-week header at row 2
    days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for i, dname in enumerate(days, start=2):
        cell = ws.cell(row=ROW_DOW, column=i, value=dname)
        cell.fill = fill_header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws.row_dimensions[ROW_DOW].height = 18

    # A/B map (anchor default = today, anchor_day default = 'B')
    if anchor_date is None:
        anchor_date = date.today()
    ab_map = ab_map_for_month_with_anchor(year, month, anchor_date, anchor_day)

    # Fairness target: total slots (weekdays * 4 periods * 3 primaries) / num_students
    wkdays = weekdays_in_month(year, month)
    total_slots = len(wkdays) * 4 * 3
    num_students = max(1, len(students))
    target_per_student = math.ceil(total_slots / num_students)

    # No "Store Stoles"; keep spacer + 4 period rows
    block_rows = [
        "8:20 - 9 AM",
        "Period 1/5",
        "Period 2/6",
        "Period 3/7 Kiosk",
        "Period 4/8",
    ]

    # Per-day assigned tracker to enforce one-per-day
    assigned_today_map: dict[date, set] = defaultdict(set)

    # Build weeks
    cur_row = ROW_DOW + 1  # start dates row at row 3
    weeks = month_weeks_sun_to_sat(year, month)
    for week in weeks:
        # Dates row (pink band)
        for col, d in enumerate(week, start=2):
            show = d.day if d.month == month else ""
            cell = ws.cell(row=cur_row, column=col, value=show)
            cell.fill = fill_pink
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_all
        cur_row += 1

        # Content rows
        for label in block_rows:
            # Left label
            lab = ws.cell(row=cur_row, column=1, value=label)
            lab.fill = fill_left
            lab.font = Font(bold=True)
            lab.alignment = Alignment(vertical="center")
            lab.border = border_all

            for col, d in enumerate(week, start=2):
                cell = ws.cell(row=cur_row, column=col)
                cell.border = border_all

                in_month = (d.month == month)
                is_weekend = d.weekday() >= 5
                ab = ab_map.get(d) if in_month else None

                # Shade weekdays by A/B; grey spillover; weekends no special shade
                if in_month and not is_weekend:
                    cell.fill = fill_day_a if ab == "A" else fill_day_b
                elif not in_month:
                    cell.fill = fill_grey

                # Do not schedule on weekends or out-of-month cells
                if (not in_month) or is_weekend:
                    continue

                # Map label to period
                if label == "8:20 - 9 AM":
                    continue  # spacer row
                if label == "Period 1/5":
                    period = 1 if ab == "A" else 5
                elif label == "Period 2/6":
                    period = 2 if ab == "A" else 6
                elif label == "Period 3/7 Kiosk":
                    period = 3 if ab == "A" else 7
                elif label == "Period 4/8":
                    period = 4 if ab == "A" else 8
                else:
                    continue

                primaries, alt = pick_for_period(
                    students, d, period,
                    assigned_today=assigned_today_map[d],
                    target_per_student=target_per_student,
                    min_gap_days=2, need=3,
                    enforce_one_per_day=True,
                    seed=seed
                )
                lines = primaries[:]
                if alt:
                    lines.append(f"Alt: {alt}")
                cell.value = "\n".join(lines) if lines else ""
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            cur_row += 1

    wb.save(out_path)


# -------------------------
# CLI
# -------------------------
def main():
    parser = argparse.ArgumentParser(
        description="CO-OP Monthly Calendar (3+Alt, fairness-balanced; A/B anchored; weekends visible but not scheduled)."
    )
    parser.add_argument("--input", default="Employee_Schedule.xlsx",
                        help="Excel with columns [Name, Schedule]")
    parser.add_argument("--out", default=None,
                        help="Output Excel filename (default: COOP_Calendar_<YYYY-MM>.xlsx)")
    # Optional overrides; if omitted, uses current month/year and today as anchor.
    parser.add_argument("--year", type=int, default=None, help="Year (defaults to current)")
    parser.add_argument("--month", type=int, default=None, help="Month 1-12 (defaults to current)")
    parser.add_argument("--anchor-day", choices=["A", "B"], default="B",
                        help="Anchor A/B day for the anchor date (default: B; e.g., Sep 3, 2025 = B)")
    parser.add_argument("--anchor-date", default=None,
                        help="Anchor date YYYY-MM-DD (defaults to today)")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    args = parser.parse_args()

    # Determine month/year from local time if not provided
    today = date.today()
    year = args.year or today.year
    month = args.month or today.month

    # Anchor so that anchor_date (default today) is args.anchor_day (default B)
    if args.anchor_date:
        anchor_date = datetime.strptime(args.anchor_date, "%Y-%m-%d").date()
    else:
        anchor_date = today

    out = args.out or f"COOP_Calendar_{year:04d}-{month:02d}.xlsx"

    students = load_students(args.input)
    build_calendar_excel(
        out_path=out,
        year=year,
        month=month,
        students=students,
        min_gap_days=2,
        seed=args.seed,
        anchor_date=anchor_date,
        anchor_day=args.anchor_day,
        enforce_one_per_day=True,
    )
    print(f"Wrote {out}")

if __name__ == "__main__":
    main()
